import openpyxl

def pad_data_in_excel(filename):
    # 打开Excel文件
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    
    # 读取第二列的数据并计算最长的长度
    data_col2 = []
    max_length = 0
    
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        col2_data = row[1]
        if col2_data:
            numbers = col2_data.split()
            data_col2.append(numbers)
            if len(numbers) > max_length:
                max_length = len(numbers)
    
    # 处理第二列数据并用0.00补充不足的部分
    for idx, numbers in enumerate(data_col2):
        while len(numbers) < max_length:
            numbers.append('0.00')
        sheet.cell(row=idx+1, column=2, value=' '.join(numbers))
    
    # 保存修改后的Excel文件
    workbook.save(filename)

# 主函数
if __name__ == "__main__":
    filename = 'current.xlsx'
    pad_data_in_excel(filename)
    print(f"{filename} 文件已修改，第二列每行长度已补充到一致。")
